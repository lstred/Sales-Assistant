"""Budget & Forecast view.

Lets the manager set per-CC growth targets and monthly seasonality weights,
then shows a full-year forecast cascaded to CC → Rep → Customer level with
current-YTD-vs-budget comparison.  Downloads to CSV or Excel.
"""

from __future__ import annotations

import logging
import os
from datetime import date, timedelta
from pathlib import Path
from typing import Callable

import pandas as pd
from PySide6.QtCore import QThread, Qt, Signal
from PySide6.QtWidgets import (
    QAbstractItemView,
    QButtonGroup,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QSpinBox,
    QSplitter,
    QTableView,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from app.config.models import AppConfig, DatabaseConfig
from app.config.store import save_config
from app.data.loaders import load_all_cost_centers, load_blended_sales, load_rep_assignments
from app.services.budget_service import (
    PERIOD_MONTH_NAMES,
    BudgetRow,
    add_ytd_actuals,
    compute_budget_by_account,
    compute_budget_by_cc,
    compute_budget_by_rep,
    parse_rep_cc_upload,
    rows_to_dataframe,
)
from app.services.fiscal_calendar import (
    build_fiscal_year,
    find_period,
    fiscal_year_for,
    fy_start_date,
)
from app.ui.theme import (
    ACCENT,
    BG,
    BORDER,
    SURFACE,
    TEXT,
    TEXT_MUTED,
)
from app.ui.views._header import ViewHeader
from app.ui.widgets.pandas_model import PandasModel

log = logging.getLogger(__name__)


# ============================================================ background loader

class _BudgetLoader(QThread):
    """Loads prior-year blended sales + assignments + CC names + current YTD."""

    loaded = Signal(
        object,   # prior_df
        object,   # curr_ytd_df
        object,   # prior_ytd_df
        object,   # assignments_df
        object,   # cc_names dict
        object,   # account_info dict
    )
    failed = Signal(str)

    def __init__(
        self,
        db: DatabaseConfig,
        budget_fy: int,
        sw: list[int],
    ) -> None:
        super().__init__()
        self._db = db
        self._budget_fy = budget_fy
        self._sw = sw

    def run(self) -> None:  # noqa: D401
        try:
            sw = self._sw
            fy = self._budget_fy
            today = date.today()

            # --- Prior full fiscal year ---
            prior_start = fy_start_date(fy - 1, sw)
            prior_end = fy_start_date(fy, sw) - timedelta(days=1)
            prior_df = load_blended_sales(
                self._db, prior_start, prior_end, None, sw, code_prefix="0"
            )

            # --- Current FY YTD through last completed period ---
            curr_start = fy_start_date(fy, sw)
            curr_ytd_df: pd.DataFrame | None = None
            prior_ytd_df: pd.DataFrame | None = None
            completed_indices: list[int] = []

            if fiscal_year_for(today) == fy:
                periods = build_fiscal_year(fy, sw)
                done = [p for p in periods if p.end < today]
                if done:
                    ytd_end = done[-1].end
                    completed_indices = [p.period - 1 for p in done]
                    curr_ytd_df = load_blended_sales(
                        self._db, curr_start, ytd_end, None, sw, code_prefix="0"
                    )
                    # Prior YTD: same period numbers in prior FY
                    done_periods = {p.period for p in done}
                    if prior_df is not None and not prior_df.empty:
                        prior_ytd_df = prior_df[
                            prior_df["fiscal_period"].isin(done_periods)
                        ]

            # --- Assignments + account info ---
            assignments_df = load_rep_assignments(self._db)
            account_info: dict[str, dict] = {}
            if assignments_df is not None and not assignments_df.empty:
                for r in assignments_df[
                    ["account_number", "old_account_number", "account_name"]
                ].drop_duplicates("account_number").itertuples(index=False):
                    acct = str(r.account_number or "").strip()
                    if acct:
                        account_info[acct] = {
                            "old": str(r.old_account_number or "").strip(),
                            "name": str(r.account_name or "").strip(),
                        }

            # --- CC names ---
            cc_df = load_all_cost_centers(self._db)
            cc_names: dict[str, str] = {}
            if cc_df is not None and not cc_df.empty:
                for r in cc_df.itertuples(index=False):
                    code = str(r.cost_center or "").strip()
                    name = str(
                        getattr(r, "cost_center_name", "") or ""
                    ).strip()
                    if code:
                        cc_names[code] = name

            self.loaded.emit(
                prior_df,
                curr_ytd_df,
                prior_ytd_df,
                assignments_df,
                cc_names,
                account_info,
            )
        except Exception as exc:  # noqa: BLE001
            log.exception("BudgetLoader failed")
            self.failed.emit(f"{type(exc).__name__}: {exc}")


# ============================================================ helpers

def _fmt_currency(v: float) -> str:
    return f"${v:,.0f}"


def _fmt_pct(v: float) -> str:
    return f"{v:.2f}%"


def _non_editable(text: str, align=Qt.AlignmentFlag.AlignRight) -> QTableWidgetItem:
    item = QTableWidgetItem(str(text))
    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
    item.setTextAlignment(align | Qt.AlignmentFlag.AlignVCenter)
    return item


def _muted(item: QTableWidgetItem) -> QTableWidgetItem:
    from PySide6.QtGui import QColor
    item.setForeground(QColor(TEXT_MUTED))
    return item


# ============================================================ settings panel

class _SettingsPanel(QWidget):
    """Left panel: budget year, CC growth table, and seasonality table."""

    compute_requested = Signal()
    saved = Signal()

    def __init__(self, cfg: AppConfig, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._cfg = cfg
        self._rep_cc_overrides: dict[tuple[str, str], float] = {}
        self.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Expanding)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        outer.addWidget(scroll, 1)

        inner = QWidget()
        scroll.setWidget(inner)
        lay = QVBoxLayout(inner)
        lay.setContentsMargins(16, 16, 16, 16)
        lay.setSpacing(16)

        # --- Budget year ---
        yr_card = QFrame()
        yr_card.setObjectName("card")
        yr_lay = QVBoxLayout(yr_card)
        yr_lay.setContentsMargins(12, 12, 12, 12)
        yr_lay.setSpacing(8)
        yr_title = QLabel("Budget Fiscal Year")
        yr_title.setStyleSheet("font-weight: 700; font-size: 13px;")
        yr_lay.addWidget(yr_title)
        yr_sub = QLabel("Year the budget is built for (prior year is the baseline).")
        yr_sub.setStyleSheet(f"color: {TEXT_MUTED}; font-size: 11px;")
        yr_sub.setWordWrap(True)
        yr_lay.addWidget(yr_sub)
        yr_row = QHBoxLayout()
        yr_row.setSpacing(8)
        self.yr_spin = QSpinBox()
        self.yr_spin.setRange(2020, 2040)
        default_fy = cfg.budget.budget_fiscal_year or fiscal_year_for(date.today())
        self.yr_spin.setValue(default_fy)
        self.yr_spin.setFixedWidth(90)
        yr_row.addWidget(self.yr_spin)
        yr_row.addStretch()
        self.compute_btn = QPushButton("Compute")
        self.compute_btn.setProperty("primary", True)
        self.compute_btn.setFixedWidth(100)
        self.compute_btn.clicked.connect(self.compute_requested)
        yr_row.addWidget(self.compute_btn)
        yr_lay.addLayout(yr_row)
        lay.addWidget(yr_card)

        # --- CC growth % table ---
        cc_card = QFrame()
        cc_card.setObjectName("card")
        cc_lay = QVBoxLayout(cc_card)
        cc_lay.setContentsMargins(12, 12, 12, 12)
        cc_lay.setSpacing(8)
        cc_title = QLabel("Cost Center Growth %")
        cc_title.setStyleSheet("font-weight: 700; font-size: 13px;")
        cc_lay.addWidget(cc_title)
        cc_sub = QLabel(
            "Enter the growth (or decline) vs prior year for each product cost center. "
            "Populated after Compute."
        )
        cc_sub.setStyleSheet(f"color: {TEXT_MUTED}; font-size: 11px;")
        cc_sub.setWordWrap(True)
        cc_lay.addWidget(cc_sub)
        self.cc_table = QTableWidget(0, 4)
        self.cc_table.setHorizontalHeaderLabels(["CC #", "CC Name", "Prior Year $", "Growth %"])
        self.cc_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.cc_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.cc_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.cc_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self.cc_table.verticalHeader().setVisible(False)
        self.cc_table.setAlternatingRowColors(True)
        self.cc_table.setMinimumHeight(180)
        self.cc_table.setMaximumHeight(320)
        self.cc_table.itemChanged.connect(self._on_cc_item_changed)
        cc_lay.addWidget(self.cc_table)
        lay.addWidget(cc_card)

        # --- Monthly seasonality table ---
        sea_card = QFrame()
        sea_card.setObjectName("card")
        sea_lay = QVBoxLayout(sea_card)
        sea_lay.setContentsMargins(12, 12, 12, 12)
        sea_lay.setSpacing(8)
        sea_title = QLabel("Monthly Seasonality %")
        sea_title.setStyleSheet("font-weight: 700; font-size: 13px;")
        sea_lay.addWidget(sea_title)
        sea_sub = QLabel(
            "Enter the % of annual sales that falls in each fiscal month "
            "(P1=Feb … P12=Jan). Values should sum to 100."
        )
        sea_sub.setStyleSheet(f"color: {TEXT_MUTED}; font-size: 11px;")
        sea_sub.setWordWrap(True)
        sea_lay.addWidget(sea_sub)
        self.sea_table = QTableWidget(12, 2)
        self.sea_table.setHorizontalHeaderLabels(["Month", "% of Year"])
        self.sea_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.sea_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.sea_table.verticalHeader().setVisible(False)
        self.sea_table.setAlternatingRowColors(True)
        self.sea_table.setFixedHeight(280)
        for i, name in enumerate(PERIOD_MONTH_NAMES):
            self.sea_table.setItem(i, 0, _non_editable(name, Qt.AlignmentFlag.AlignLeft))
            pct = cfg.budget.monthly_seasonality_pct[i] if i < len(cfg.budget.monthly_seasonality_pct) else 8.33
            pct_item = QTableWidgetItem(f"{pct:.2f}")
            pct_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.sea_table.setItem(i, 1, pct_item)
        self.sea_table.itemChanged.connect(self._on_sea_item_changed)
        sea_lay.addWidget(self.sea_table)
        self.sea_total_label = QLabel("Total: 0.00%")
        self.sea_total_label.setStyleSheet(f"font-size: 11px; color: {TEXT_MUTED};")
        sea_lay.addWidget(self.sea_total_label)
        self._update_sea_total()
        lay.addWidget(sea_card)

        # --- Save button ---
        self.save_btn = QPushButton("Save Settings")
        self.save_btn.clicked.connect(self._save)
        lay.addWidget(self.save_btn)

        # --- Rep-CC growth upload card ---
        up_card = QFrame()
        up_card.setObjectName("card")
        up_lay = QVBoxLayout(up_card)
        up_lay.setContentsMargins(12, 12, 12, 12)
        up_lay.setSpacing(8)
        up_title = QLabel("Rep-Level Growth Override (Upload)")
        up_title.setStyleSheet("font-weight: 700; font-size: 13px;")
        up_lay.addWidget(up_title)
        up_sub = QLabel(
            "Upload a CSV or Excel file to set a unique growth % per rep per cost center. "
            "Overrides the CC-level table above for matching rep+CC pairs. "
            "Missing combinations fall back to the CC-level %."
        )
        up_sub.setStyleSheet(f"color: {TEXT_MUTED}; font-size: 11px;")
        up_sub.setWordWrap(True)
        up_lay.addWidget(up_sub)

        # Format spec box
        spec = QLabel(
            "<b>Required columns (exact names, case-insensitive):</b><br>"
            "&nbsp;&nbsp;<b>rep_number</b> — salesman number (e.g. <tt>42</tt>)<br>"
            "&nbsp;&nbsp;<b>cost_center</b> — CC code (e.g. <tt>010</tt>)<br>"
            "&nbsp;&nbsp;<b>growth_pct</b> — growth or decline % (<tt>10</tt> = +10 %, <tt>-5</tt> = −5 %)<br>"
            "One row per rep × CC pair. Extra columns are ignored."
        )
        spec.setStyleSheet(
            f"background: #EFF6FF; border: 1px solid #BFDBFE; border-radius: 6px; "
            f"padding: 8px; font-size: 11px; color: {TEXT};"
        )
        spec.setWordWrap(True)
        up_lay.addWidget(spec)

        up_btn_row = QHBoxLayout()
        up_btn_row.setSpacing(6)
        self._upload_btn = QPushButton("⬆  Upload CSV / Excel")
        self._upload_btn.clicked.connect(self._upload_overrides)
        up_btn_row.addWidget(self._upload_btn, 1)
        self._dl_template_btn = QPushButton("⬇  Template")
        self._dl_template_btn.clicked.connect(self._download_template)
        up_btn_row.addWidget(self._dl_template_btn)
        up_lay.addLayout(up_btn_row)

        self._upload_status = QLabel("No file loaded — using CC-level growth % for all reps.")
        self._upload_status.setStyleSheet(f"color: {TEXT_MUTED}; font-size: 11px;")
        self._upload_status.setWordWrap(True)
        up_lay.addWidget(self._upload_status)

        self._upload_preview = QTableWidget(0, 3)
        self._upload_preview.setHorizontalHeaderLabels(["Rep #", "CC", "Growth %"])
        self._upload_preview.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self._upload_preview.verticalHeader().setVisible(False)
        self._upload_preview.setAlternatingRowColors(True)
        self._upload_preview.setMaximumHeight(140)
        self._upload_preview.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._upload_preview.setVisible(False)
        up_lay.addWidget(self._upload_preview)

        lay.addWidget(up_card)
        lay.addStretch()

    # ------------------------------------------------------------ public
    def budget_year(self) -> int:
        return self.yr_spin.value()

    def growth_pct_map(self) -> dict[str, float]:
        """Current CC → growth % from the table."""
        out: dict[str, float] = {}
        for row in range(self.cc_table.rowCount()):
            cc_item = self.cc_table.item(row, 0)
            pct_item = self.cc_table.item(row, 3)
            if cc_item and pct_item:
                cc = cc_item.text().strip()
                try:
                    out[cc] = float(pct_item.text().replace("%", "").strip())
                except ValueError:
                    out[cc] = 0.0
        return out

    def rep_cc_growth_pct(self) -> dict[tuple[str, str], float]:
        """Rep-CC override map from uploaded file (empty if no file loaded)."""
        return dict(self._rep_cc_overrides)

    def seasonality_pct(self) -> list[float]:
        result = []
        for i in range(self.sea_table.rowCount()):
            item = self.sea_table.item(i, 1)
            try:
                result.append(float(item.text().replace("%", "").strip()) if item else 8.33)
            except ValueError:
                result.append(8.33)
        return result

    def populate_cc_table(
        self,
        cc_rows: list[tuple[str, str, float]],  # (cc_code, cc_name, prior_sales)
        saved_growth: dict[str, float],
    ) -> None:
        """Populate the CC growth % table from computed prior-year data."""
        self.cc_table.blockSignals(True)
        self.cc_table.setRowCount(len(cc_rows))
        for i, (code, name, prior) in enumerate(cc_rows):
            self.cc_table.setItem(i, 0, _non_editable(code, Qt.AlignmentFlag.AlignLeft))
            self.cc_table.setItem(i, 1, _non_editable(name, Qt.AlignmentFlag.AlignLeft))
            self.cc_table.setItem(i, 2, _non_editable(_fmt_currency(prior)))
            pct = saved_growth.get(code, 0.0)
            pct_item = QTableWidgetItem(f"{pct:.2f}")
            pct_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.cc_table.setItem(i, 3, pct_item)
        self.cc_table.blockSignals(False)

    # ------------------------------------------------------------ private
    def _on_cc_item_changed(self, item: QTableWidgetItem) -> None:
        if item.column() != 3:
            return
        # Validate numeric
        try:
            float(item.text().replace("%", "").strip())
        except ValueError:
            item.setText("0.00")

    # ------------------------------------------------------------ upload
    def _upload_overrides(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Open Rep-CC Growth Override File",
            str(Path.home()),
            "Spreadsheets (*.csv *.xlsx *.xls)",
        )
        if not path:
            return
        overrides, errors = parse_rep_cc_upload(path)
        if errors:
            detail = "\n".join(errors[:10])
            if len(errors) > 10:
                detail += f"\n… and {len(errors) - 10} more."
            reply = QMessageBox.warning(
                self, "Upload warnings",
                f"{len(overrides)} overrides loaded with {len(errors)} warning(s):\n\n{detail}",
                QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel,
            )
            if reply == QMessageBox.StandardButton.Cancel:
                return
        if not overrides and not errors:
            QMessageBox.information(self, "Empty file", "No valid rows found in the file.")
            return
        self._rep_cc_overrides = overrides
        self._refresh_upload_preview()
        if overrides:
            self._upload_status.setText(
                f"{len(overrides)} rep × CC override(s) loaded from {os.path.basename(path)}. "
                "CC-level % used as fallback for any unspecified rep × CC pairs."
            )
            self._upload_status.setStyleSheet("font-size: 11px; font-weight: 600; color: #16A34A;")
        else:
            self._upload_status.setText("Upload produced no usable rows — check the file.")
            self._upload_status.setStyleSheet(f"font-size: 11px; color: #DC2626;")

    def _refresh_upload_preview(self) -> None:
        data = sorted(self._rep_cc_overrides.items())
        self._upload_preview.setRowCount(len(data))
        for i, ((rep, cc), pct) in enumerate(data):
            self._upload_preview.setItem(i, 0, _non_editable(rep, Qt.AlignmentFlag.AlignLeft))
            self._upload_preview.setItem(i, 1, _non_editable(cc, Qt.AlignmentFlag.AlignLeft))
            self._upload_preview.setItem(i, 2, _non_editable(f"{pct:+.2f}%"))
        self._upload_preview.setVisible(bool(data))

    def _download_template(self) -> None:
        """Save a blank template CSV so the manager knows the exact format."""
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Upload Template",
            str(Path.home() / "budget_override_template.csv"),
            "CSV files (*.csv)",
        )
        if not path:
            return
        template = (
            "rep_number,cost_center,growth_pct\n"
            "# Instructions:\n"
            "#   rep_number  — salesman number exactly as it appears in Sales Reps (e.g. 42)\n"
            "#   cost_center — product CC code (e.g. 010)\n"
            "#   growth_pct  — numeric % (e.g. 10 for +10%, -5 for -5%)\n"
            "#   Remove these comment lines before uploading.\n"
            "42,010,10.0\n"
            "42,020,-5.0\n"
            "17,010,12.5\n"
        )
        try:
            Path(path).write_text(template)
            self._upload_status.setText(f"Template saved: {os.path.basename(path)}")
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "Save failed", str(exc))

    def _on_sea_item_changed(self, item: QTableWidgetItem) -> None:
        if item.column() != 1:
            return
        try:
            float(item.text().replace("%", "").strip())
        except ValueError:
            item.setText("8.33")
        self._update_sea_total()

    def _update_sea_total(self) -> None:
        total = sum(self.seasonality_pct())
        color = "#16A34A" if abs(total - 100.0) < 0.05 else "#DC2626"
        self.sea_total_label.setText(f"Total: {total:.2f}%")
        self.sea_total_label.setStyleSheet(f"font-size: 11px; font-weight: 600; color: {color};")

    def _save(self) -> None:
        self._cfg.budget.budget_fiscal_year = self.yr_spin.value()
        self._cfg.budget.cc_growth_pct = self.growth_pct_map()
        self._cfg.budget.monthly_seasonality_pct = self.seasonality_pct()
        try:
            save_config(self._cfg)
            self.saved.emit()
        except Exception as exc:  # noqa: BLE001
            log.exception("save_config failed in budget view")
            QMessageBox.warning(self, "Save Failed", str(exc))


# ============================================================ main view

class BudgetView(QWidget):
    """Budget & Forecast view — main entry point."""

    def __init__(
        self,
        cfg: AppConfig,
        get_db: Callable[[], DatabaseConfig],
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self._cfg = cfg
        self._get_db = get_db

        # State
        self._prior_df: pd.DataFrame | None = None
        self._curr_ytd_df: pd.DataFrame | None = None
        self._prior_ytd_df: pd.DataFrame | None = None
        self._assignments_df: pd.DataFrame | None = None
        self._cc_names: dict[str, str] = {}
        self._account_info: dict[str, dict] = {}
        self._completed_indices: list[int] = []
        self._rows_by_cc: list[BudgetRow] = []
        self._rows_by_rep: list[BudgetRow] = []
        self._rows_by_acct: list[BudgetRow] = []
        self._current_mode = "cc"
        self._loaders: list[_BudgetLoader] = []

        root = QVBoxLayout(self)
        root.setContentsMargins(28, 24, 28, 24)
        root.setSpacing(12)

        root.addWidget(
            ViewHeader(
                "Budget & Forecast",
                "Set per-CC growth targets and monthly seasonality, then export "
                "full-year forecasts by cost center, sales rep, or customer.",
            )
        )

        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setChildrenCollapsible(False)

        # Left: settings panel
        self._settings = _SettingsPanel(cfg)
        self._settings.setMinimumWidth(340)
        self._settings.setMaximumWidth(440)
        self._settings.compute_requested.connect(self._on_compute)
        self._settings.saved.connect(self._on_saved)
        splitter.addWidget(self._settings)

        # Right: results panel
        results_widget = QWidget()
        results_lay = QVBoxLayout(results_widget)
        results_lay.setContentsMargins(0, 0, 0, 0)
        results_lay.setSpacing(10)

        # Mode toggle + download bar
        top_bar = QHBoxLayout()
        top_bar.setSpacing(6)

        mode_label = QLabel("View by:")
        mode_label.setStyleSheet(f"color: {TEXT_MUTED}; font-size: 12px;")
        top_bar.addWidget(mode_label)

        self._mode_group = QButtonGroup(self)
        self._mode_group.setExclusive(True)
        for key, label in (("cc", "Cost Center"), ("rep", "Sales Rep"), ("account", "Customer")):
            btn = QPushButton(label)
            btn.setCheckable(True)
            btn.setChecked(key == "cc")
            self._mode_group.addButton(btn)
            btn.clicked.connect(lambda _=False, k=key: self._switch_mode(k))
            top_bar.addWidget(btn)

        top_bar.addStretch()

        dl_csv_btn = QPushButton("⬇  Download CSV")
        dl_csv_btn.clicked.connect(self._download_csv)
        top_bar.addWidget(dl_csv_btn)

        dl_xl_btn = QPushButton("⬇  Download Excel")
        dl_xl_btn.clicked.connect(self._download_excel)
        top_bar.addWidget(dl_xl_btn)

        results_lay.addLayout(top_bar)

        # Status bar
        self._status_label = QLabel(
            "Click Compute to load prior-year data and build the forecast."
        )
        self._status_label.setStyleSheet(f"color: {TEXT_MUTED}; font-size: 12px;")
        results_lay.addWidget(self._status_label)

        # Results table
        self._model = PandasModel()
        self._table = QTableView()
        self._table.setModel(self._model)
        self._table.setAlternatingRowColors(True)
        self._table.verticalHeader().setVisible(False)
        self._table.setSortingEnabled(True)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        results_lay.addWidget(self._table, 1)

        splitter.addWidget(results_widget)
        splitter.setSizes([380, 900])

        root.addWidget(splitter, 1)

    # ------------------------------------------------------------ slots

    def _on_compute(self) -> None:
        fy = self._settings.budget_year()
        sw = list(self._cfg.fiscal.six_week_january_years)
        self._status_label.setText(
            f"Loading prior-year data for FY{fy - 1} and current YTD…"
        )
        self._settings.compute_btn.setEnabled(False)

        loader = _BudgetLoader(self._get_db(), fy, sw)
        loader.loaded.connect(self._on_loaded)
        loader.failed.connect(self._on_load_failed)
        self._loaders.append(loader)
        loader.finished.connect(
            lambda ldr=loader: self._loaders.remove(ldr) if ldr in self._loaders else None
        )
        loader.start()

    def _on_loaded(
        self,
        prior_df,
        curr_ytd_df,
        prior_ytd_df,
        assignments_df,
        cc_names,
        account_info,
    ) -> None:
        self._prior_df = prior_df
        self._curr_ytd_df = curr_ytd_df
        self._prior_ytd_df = prior_ytd_df
        self._assignments_df = assignments_df
        self._cc_names = cc_names or {}
        self._account_info = account_info or {}

        # Determine completed period indices for the budget year
        fy = self._settings.budget_year()
        sw = list(self._cfg.fiscal.six_week_january_years)
        today = date.today()
        from app.services.fiscal_calendar import fiscal_year_for
        if fiscal_year_for(today) == fy:
            periods = __import__("app.services.fiscal_calendar", fromlist=["build_fiscal_year"]).build_fiscal_year(fy, sw)
            self._completed_indices = [p.period - 1 for p in periods if p.end < today]
        else:
            self._completed_indices = []

        # Populate CC table from prior data
        cc_prior: dict[str, float] = {}
        if prior_df is not None and not prior_df.empty:
            grouped = prior_df.groupby("cost_center")["revenue"].sum()
            cc_prior = grouped.to_dict()

        # All product CCs (starting with '0')
        all_product_ccs = sorted(
            {cc for cc in (list(cc_prior.keys()) + list(cc_names.keys())) if str(cc).startswith("0")}
        )
        cc_rows = [
            (cc, cc_names.get(cc, cc), float(cc_prior.get(cc, 0.0)))
            for cc in all_product_ccs
        ]
        self._settings.populate_cc_table(cc_rows, self._cfg.budget.cc_growth_pct)

        # Compute budget rows
        self._recompute()

        prior_lines = len(prior_df) if prior_df is not None else 0
        ytd_lines = len(curr_ytd_df) if curr_ytd_df is not None else 0
        self._status_label.setText(
            f"FY{fy - 1} prior: {prior_lines:,} lines  ·  "
            f"Current YTD: {ytd_lines:,} lines  ·  "
            f"Completed periods: {len(self._completed_indices)}"
        )
        self._settings.compute_btn.setEnabled(True)

    def _on_load_failed(self, msg: str) -> None:
        self._settings.compute_btn.setEnabled(True)
        self._status_label.setText(f"Load failed — {msg}")
        log.error("BudgetLoader error: %s", msg)

    def _on_saved(self) -> None:
        self._status_label.setText("Settings saved.")

    def _switch_mode(self, mode: str) -> None:
        self._current_mode = mode
        self._refresh_table()

    # ------------------------------------------------------------ compute

    def _recompute(self) -> None:
        growth = self._settings.growth_pct_map()
        seasonality = self._settings.seasonality_pct()
        rep_cc = self._settings.rep_cc_growth_pct() or None  # None = no overrides

        self._rows_by_cc = compute_budget_by_cc(
            self._prior_df, growth, seasonality, self._cc_names,
            rep_cc_growth_pct=rep_cc,
            assignments_df=self._assignments_df,
        )
        self._rows_by_rep = compute_budget_by_rep(
            self._prior_df, self._assignments_df, growth, seasonality, self._cc_names,
            rep_cc_growth_pct=rep_cc,
        )
        self._rows_by_acct = compute_budget_by_account(
            self._prior_df, self._assignments_df, self._account_info,
            growth, seasonality, self._cc_names,
            rep_cc_growth_pct=rep_cc,
        )

        for rows, gby in (
            (self._rows_by_cc, "cc"),
            (self._rows_by_rep, "rep"),
            (self._rows_by_acct, "account"),
        ):
            add_ytd_actuals(
                rows,
                self._curr_ytd_df,
                self._prior_ytd_df,
                self._completed_indices,
                group_by=gby,
            )

        self._refresh_table()

    def _refresh_table(self) -> None:
        rows = {
            "cc": self._rows_by_cc,
            "rep": self._rows_by_rep,
            "account": self._rows_by_acct,
        }.get(self._current_mode, [])

        has_ytd = bool(self._completed_indices)
        df = rows_to_dataframe(
            rows,
            mode=self._current_mode,
            include_monthly=False,
            include_ytd=has_ytd,
        )

        # Format currency/pct columns for display
        for col in df.columns:
            if col.endswith("$") or "Sales" in col or "Budget" in col or "Actual" in col or "Change" in col:
                try:
                    df[col] = df[col].apply(lambda v: _fmt_currency(float(v)) if pd.notna(v) else "")
                except Exception:  # noqa: BLE001
                    pass
            if col == "Growth %":
                try:
                    df[col] = df[col].apply(lambda v: _fmt_pct(float(v)) if pd.notna(v) else "")
                except Exception:  # noqa: BLE001
                    pass

        self._model.set_dataframe(df)
        self._table.resizeColumnsToContents()

    # ------------------------------------------------------------ downloads

    def _current_export_df(self, mode: str) -> pd.DataFrame:
        rows = {
            "cc": self._rows_by_cc,
            "rep": self._rows_by_rep,
            "account": self._rows_by_acct,
        }.get(mode, [])
        return rows_to_dataframe(rows, mode=mode, include_monthly=True, include_ytd=False)

    def _download_csv(self) -> None:
        if not self._rows_by_cc:
            QMessageBox.information(self, "No Data", "Click Compute first.")
            return
        mode, label = self._pick_export_mode()
        if mode is None:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, f"Save CSV — By {label}",
            str(Path.home() / f"budget_by_{mode}.csv"),
            "CSV files (*.csv)",
        )
        if not path:
            return
        try:
            self._current_export_df(mode).to_csv(path, index=False)
            self._status_label.setText(f"Saved: {os.path.basename(path)}")
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "Export Failed", str(exc))

    def _download_excel(self) -> None:
        if not self._rows_by_cc:
            QMessageBox.information(self, "No Data", "Click Compute first.")
            return
        mode, label = self._pick_export_mode()
        if mode is None:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, f"Save Excel — By {label}",
            str(Path.home() / f"budget_by_{mode}.xlsx"),
            "Excel files (*.xlsx)",
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter

            df = self._current_export_df(mode)
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name=f"Budget by {label}")
                ws = writer.sheets[f"Budget by {label}"]

                # Header styling
                hdr_fill = PatternFill("solid", fgColor="0F172A")
                hdr_font = Font(bold=True, color="F8FAFC", size=11)
                for cell in ws[1]:
                    cell.fill = hdr_fill
                    cell.font = hdr_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Auto-width
                for col_idx, col_name in enumerate(df.columns, 1):
                    col_letter = get_column_letter(col_idx)
                    max_len = max(
                        len(str(col_name)),
                        *(len(str(v)) for v in df.iloc[:, col_idx - 1].fillna(""))
                    )
                    ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

                # Currency columns
                currency_cols = [
                    i + 1 for i, c in enumerate(df.columns)
                    if any(kw in c for kw in ("Sales", "Budget", "Change", "Actual"))
                ]
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        if cell.column in currency_cols:
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal="right")

            self._status_label.setText(f"Saved: {os.path.basename(path)}")
        except ImportError:
            QMessageBox.warning(
                self, "Missing Dependency",
                "openpyxl is required for Excel export.\nRun: pip install openpyxl"
            )
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "Export Failed", str(exc))

    def _pick_export_mode(self) -> tuple[str | None, str]:
        """Show a quick dialog to pick export grouping level."""
        from PySide6.QtWidgets import QDialog, QDialogButtonBox
        dlg = QDialog(self)
        dlg.setWindowTitle("Export by…")
        dlg.setFixedWidth(280)
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(16, 16, 16, 16)
        lay.setSpacing(10)
        lay.addWidget(QLabel("Choose export grouping:"))
        grp = QButtonGroup(dlg)
        grp.setExclusive(True)
        btns: dict[str, QPushButton] = {}
        for key, lbl in (("cc", "Cost Center"), ("rep", "Sales Rep"), ("account", "Customer")):
            btn = QPushButton(lbl)
            btn.setCheckable(True)
            btn.setChecked(key == self._current_mode)
            grp.addButton(btn)
            lay.addWidget(btn)
            btns[key] = btn
        bb = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        bb.accepted.connect(dlg.accept)
        bb.rejected.connect(dlg.reject)
        lay.addWidget(bb)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return None, ""
        for key, btn in btns.items():
            if btn.isChecked():
                labels = {"cc": "Cost Center", "rep": "Sales Rep", "account": "Customer"}
                return key, labels[key]
        return self._current_mode, "Cost Center"
